import sqlite3
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

# SQLite 데이터베이스와 연결하는 함수
def connect_db():
    return sqlite3.connect('restaurant_menu.db')

# SQLite 데이터베이스 초기화 함수
def reset_database():
    conn = sqlite3.connect('restaurant_menu.db')
    c = conn.cursor()

    # 기존 테이블 삭제 (데이터베이스를 초기화하기 위해)
    c.execute('DROP TABLE IF EXISTS MenuIngredients')
    c.execute('DROP TABLE IF EXISTS Menus')
    c.execute('DROP TABLE IF EXISTS Ingredients')

    conn.commit()
    conn.close()

    # 데이터베이스 다시 생성
    create_database()
    insert_data()

# SQLite 데이터베이스에 유닛를 삽입하는 함수
def insert_menu(menu_name):
    conn = connect_db()
    c = conn.cursor()
    c.execute('INSERT INTO Menus (MenuName) VALUES (?)', (menu_name,))
    conn.commit()
    conn.close()

# SQLite 데이터베이스에 부품를 삽입하는 함수
def insert_ingredient(ingredient_name, price):
    conn = connect_db()
    c = conn.cursor()
    c.execute('INSERT INTO Ingredients (IngredientName, Price) VALUES (?, ?)', (ingredient_name, price))
    conn.commit()
    conn.close()

# SQLite 데이터베이스에서 유닛와 부품를 조회하는 함수
def get_data():
    conn = connect_db()
    c = conn.cursor()
    try:
        c.execute('SELECT * FROM Menus')
        menus = c.fetchall()
        c.execute('SELECT * FROM Ingredients')
        ingredients = c.fetchall()
        c.execute('''SELECT mi.MenuID, m.MenuName, mi.IngredientID, i.IngredientName, mi.Quantity, i.Price
                     FROM MenuIngredients mi
                     JOIN Menus m ON mi.MenuID = m.MenuID
                     JOIN Ingredients i ON mi.IngredientID = i.IngredientID''')
        menu_ingredients = c.fetchall()
        conn.close()
        return menus, ingredients, menu_ingredients
    except sqlite3.Error as e:
        st.error(f"SQLite error: {e}")
        return [], [], []  # 에러 발생 시 빈 리스트 반환

def create_database():
    conn = sqlite3.connect('restaurant_menu.db')
    c = conn.cursor()

    # Menus 테이블 생성
    c.execute('''CREATE TABLE IF NOT EXISTS Menus (
                    MenuID INTEGER PRIMARY KEY AUTOINCREMENT,
                    MenuName TEXT NOT NULL,
                    DateSubmitted TEXT NOT NULL
                )''')

    # Ingredients 테이블 생성
    c.execute('''CREATE TABLE IF NOT EXISTS Ingredients (
                    IngredientID INTEGER PRIMARY KEY AUTOINCREMENT,
                    IngredientName TEXT NOT NULL,
                    Price REAL NOT NULL
                )''')

    # MenuIngredients 테이블 생성
    c.execute('''CREATE TABLE IF NOT EXISTS MenuIngredients (
                    MenuID INTEGER,
                    IngredientID INTEGER,
                    Quantity REAL NOT NULL,
                    FOREIGN KEY (MenuID) REFERENCES Menus(MenuID),
                    FOREIGN KEY (IngredientID) REFERENCES Ingredients(IngredientID)
                )''')

    conn.commit()
    conn.close()

def insert_data():
    conn = sqlite3.connect('restaurant_menu.db')
    c = conn.cursor()

    # 유닛 데이터 추가
    menus = [
        ('Microcontrollers and Processors', '2024-01-02'),
        ('Memory and Storage', '2024-01-03'),
        ('Power Management', '2024-01-04'),
        ('Oscillators and Timing', '2024-01-05'),
        ('Communication Modules', '2024-01-06'),
        ('Input/Output Interfaces', '2024-01-07'),
        ('Connectors and Interfaces', '2024-01-08'),
        ('Displays and Indicators', '2024-01-09'),
        ('Passive Components', '2024-01-10'),
        ('Mechanical and Assembly Components', '2024-01-11'),
        ('Control and Feedback', '2024-01-12')
    ]
    c.executemany('INSERT INTO Menus (MenuName, DateSubmitted) VALUES (?, ?)', menus)

    # 부품 데이터 추가
    ingredients = [
        ('ATmega328', 2.0),
        ('STM32', 0.5),
        ('ESP32', 0.3),
        ('EEPROM', 0.2),
        ('Oscillator', 0.1),
        ('Voltage Regulator', 0.1),
        ('DC-DC Converter', 0.2),
        ('Pin Header', 0.05),
        ('JST Connector', 0.05),
        ('USB Port', 0.1),
        ('UART Module', 0.1),
        ('I2C Module', 0.1),
        ('SPI Module', 0.1),
        ('Wi-Fi Module', 0.2),
        ('Bluetooth Module', 0.2),
        ('JTAG', 0.1),
        ('SWD (Serial Wire Debug)', 0.1),
        ('GPIO Ports', 0.2),
        ('ADC (Analog-to-Digital Converter)', 0.2),
        ('DAC (Digital-to-Analog Converter)', 0.2),
        ('LED Indicators', 0.05),
        ('Push Button Switches', 0.05),
        ('PCB', 0.3),
        ('Heat Sink', 0.1),
        ('Resistors', 0.1),
        ('Capacitors', 0.1),
        ('Diodes', 0.1),
        ('Transistors', 0.1),
        ('Inductors', 0.1),
        ('Relays', 0.2),
        ('Potentiometers', 0.1),
        ('Temperature Sensors', 0.2),
        ('Hall Effect Sensors', 0.2),
        ('Touch Sensors', 0.2),
        ('Relay Module', 0.2),
        ('Buzzer', 0.1),
        ('RTC (Real-Time Clock)', 0.2),
        ('LCD Display', 0.3),
        ('OLED Display', 0.3),
        ('Battery Holder', 0.1),
        ('Fuse', 0.05),
    ]
    c.executemany('INSERT INTO Ingredients (IngredientName, Price) VALUES (?, ?)', ingredients)

    # 유닛-부품 관계 데이터 추가
    menu_ingredients = [
        (1, 1, 2.0), (1, 2, 0.5), (1, 3, 0.3),
        (2, 4, 0.2),
        (3, 5, 0.1), (3, 6, 0.2), (3, 7, 0.1),
        (4, 8, 0.1), (4, 9, 0.2),
        (5, 10, 0.1), (5, 11, 0.1), (5, 12, 0.1), (5, 13, 0.2), (5, 14, 0.2),
        (6, 15, 0.2), (6, 16, 0.2), (6, 17, 0.2), (6, 18, 0.2), (6, 19, 0.2), (6, 20, 0.2),
        (7, 21, 0.05), (7, 22, 0.05), (7, 23, 0.1), (7, 24, 0.2),
        (8, 25, 0.05), (8, 26, 0.3), (8, 27, 0.3),
        (9, 28, 0.1), (9, 29, 0.1), (9, 30, 0.1), (9, 31, 0.1), (9, 32, 0.1),
        (10, 33, 0.3), (10, 34, 0.1), (10, 35, 0.05),
        (11, 36, 0.05), (11, 37, 0.1), (11, 38, 0.2)
    ]
    c.executemany('INSERT INTO MenuIngredients (MenuID, IngredientID, Quantity) VALUES (?, ?, ?)', menu_ingredients)

    conn.commit()
    conn.close()


# 유닛 총 비용 계산 함수
def calculate_menu_costs(menu_df, menu_ingredients_df, ingredient_df):
    menu_costs = []
    for _, menu_row in menu_df.iterrows():
        menu_id = menu_row["ID"]
        quantity = menu_row["Quantity"]
        total_cost = 0

        # 유닛에 포함된 부품를 찾아서 총 비용 계산
        menu_ingredients = menu_ingredients_df[menu_ingredients_df["MenuID"] == menu_id]
        for _, mi_row in menu_ingredients.iterrows():
            ingredient_id = mi_row["IngredientID"]
            ingredient_price = ingredient_df[ingredient_df["ID"] == ingredient_id]["Price"].values[0]
            ingredient_quantity = mi_row["Quantity"] * quantity
            total_cost += ingredient_price * ingredient_quantity

        menu_costs.append(total_cost)
    
    return menu_costs

def main():
    st.set_page_config(page_title="자동 견적")
    st.title("GSI 프로젝트 및 부품관리 시스템")

    if st.sidebar.button("초기화"):
        reset_database()
        st.sidebar.success("데이터베이스가 초기화되었습니다!")

    if "menu_df" not in st.session_state or "ingredient_df" not in st.session_state:
        menus, ingredients, menu_ingredients = get_data()

        # 필요 시 필요한 열만 선택하여 DataFrame 생성
        if menus and len(menus[0]) == 3:
            st.session_state.menu_df = pd.DataFrame(menus, columns=["ID", "MenuName", "DateSubmitted"]).drop(columns=["DateSubmitted"])
        else:
            st.session_state.menu_df = pd.DataFrame(menus, columns=["ID", "MenuName"])

        st.session_state.ingredient_df = pd.DataFrame(ingredients, columns=["ID", "IngredientName", "Price"])
        st.session_state.menu_ingredients_df = pd.DataFrame(menu_ingredients, columns=["MenuID", "MenuName", "IngredientID", "IngredientName", "Quantity", "Price"])
        st.session_state.new_df=pd.DataFrame(ingredients, columns=["ID", "IngredientName", "Price"])
        
        # 유닛별 수량 및 총 비용 초기화
        st.session_state.menu_df["Quantity"] = 0
        st.session_state.menu_df["TotalCost"] = 0

    st.sidebar.header("부품 추가")
    with st.sidebar.form("부품 추가 양식"):
        ingredient_name = st.text_area("부품 이름")
        price = st.number_input("가격", min_value=0.0, step=0.1)
        ingredient_submitted = st.form_submit_button("부품 추가")

    if ingredient_submitted:
        insert_ingredient(ingredient_name, price)
        st.sidebar.write("부품이 추가되었습니다!")
        _, ingredients, _ = get_data()
        st.session_state.ingredient_df = pd.DataFrame(ingredients, columns=["ID", "IngredientName", "Price"])

    st.sidebar.header("유닛 추가")
    with st.sidebar.form("유닛 추가 양식"):
        menu_name = st.text_area("유닛 이름")
        
        # 부품 선택 및 수량 입력
        ingredients_list = st.session_state.ingredient_df[['ID', 'IngredientName']].to_dict(orient='records')
        ingredient_options = {ingredient['IngredientName']: ingredient['ID'] for ingredient in ingredients_list}
        
        selected_ingredients = st.multiselect("부품 선택", options=list(ingredient_options.keys()))
        quantities = {ingredient: st.number_input(f"{ingredient} 수량", min_value=0, step=1, key=f"quantity_{ingredient_options[ingredient]}") for ingredient in selected_ingredients}
        
        menu_submitted = st.form_submit_button("유닛 추가")

        if menu_submitted:
            conn = connect_db()
            c = conn.cursor()
            c.execute('INSERT INTO Menus (MenuName, DateSubmitted) VALUES (?, ?)', (menu_name, '2024-08-10'))
            menu_id = c.lastrowid

            # 선택된 부품와 수량을 MenuIngredients 테이블에 추가
            for ingredient_name in selected_ingredients:
                ingredient_id = ingredient_options[ingredient_name]
                quantity = quantities[ingredient_name]
                c.execute('INSERT INTO MenuIngredients (MenuID, IngredientID, Quantity) VALUES (?, ?, ?)', (menu_id, ingredient_id, quantity))

            conn.commit()
            conn.close()
            st.sidebar.write("유닛와 부품가 추가되었습니다!")
            menus, _, _ = get_data()
            st.session_state.menu_df = pd.DataFrame(menus, columns=["ID", "MenuName", "DateSubmitted"]).drop(columns=["DateSubmitted"])
            st.session_state.menu_df["Quantity"] = 0
            st.session_state.menu_df["TotalCost"] = 0

    # 유닛 선택 및 수량 입력
    selected_menu = st.multiselect("유닛 선택", st.session_state.menu_df["MenuName"].tolist())
    st.write("선택한 유닛:")
    for menu_name in selected_menu:
        menu_id = st.session_state.menu_df[st.session_state.menu_df["MenuName"] == menu_name]["ID"].values[0]
        quantity_key = f"quantity_{menu_id}"  # 고유 키 생성
        quantity = st.number_input(f"{menu_name} 수량", min_value=0, step=1, key=quantity_key)
        st.session_state.menu_df.loc[st.session_state.menu_df["ID"] == menu_id, "Quantity"] = quantity

    # 선택한 유닛의 부품를 조회하고 업데이트된 부품 테이블에 표시
    filtered_ingredients = st.session_state.menu_ingredients_df[st.session_state.menu_ingredients_df["MenuName"].isin(selected_menu)]
    filtered_ingredient_ids = filtered_ingredients["IngredientID"].unique()
    updated_ingredients = st.session_state.ingredient_df[st.session_state.ingredient_df["ID"].isin(filtered_ingredient_ids)].copy()
    updated_ingredients["Include"] = False  # 체크박스 열 추가

    if "Quantity" not in st.session_state.ingredient_df.columns:
        st.session_state.ingredient_df["Quantity"] = 0  # 기본값 설정

    # Quantity는 수정 가능하게 설정
    edited_updated_ingredients = st.data_editor(
        updated_ingredients,
        num_rows="dynamic",
        column_config={
            "Quantity": st.column_config.NumberColumn("수량"),
            "Include": st.column_config.CheckboxColumn("Include")
        }
    )

    if edited_updated_ingredients is not None:
        st.session_state.ingredient_df.update(edited_updated_ingredients)
        st.session_state.ingredient_df["TotalCost"] = st.session_state.ingredient_df["Price"] * st.session_state.ingredient_df["Quantity"]
    
    st.subheader("부품")
    # 새로운 열이 있는지 확인하고 없으면 추가
    if "Quantity" not in st.session_state.new_df.columns:
        st.session_state.new_df["Quantity"] = 0  # 수량을 입력할 수 있는 열 추가
    if "TotalCost" not in st.session_state.new_df.columns:
        st.session_state.new_df["TotalCost"] = 0  # 총 비용을 계산할 열 추가
    if "Include" not in st.session_state.new_df.columns:
        st.session_state.new_df["Include"] = False  # 체크박스 열 추가

    # 'Include' 열이 없으면 추가
    if "Include" not in st.session_state.ingredient_df.columns:
        st.session_state.ingredient_df["Include"] = False

    # 새로운 열이 있는지 확인하고 없으면 추가
    if "Include" not in st.session_state.new_df.columns:
        st.session_state.new_df["Include"] = False  # 체크박스 열 추가

    # Quantity와 Include 열을 편집할 수 있도록 설정
    edited_new_df = st.data_editor(
        st.session_state.new_df,
        num_rows="dynamic",
        column_config={
            "Quantity": st.column_config.NumberColumn("수량", min_value=0, step=1),
            "Include": st.column_config.CheckboxColumn("Include")
        }
    )

    # 편집된 데이터가 있는 경우 업데이트
    if edited_new_df is not None:
        st.session_state.new_df.update(edited_new_df)
        st.session_state.new_df["TotalCost"] = st.session_state.new_df["Price"] * st.session_state.new_df["Quantity"]

    # 기존 부품와 새 부품 모두 'Include' 열로 필터링
    selected_ingredients = pd.concat([
        st.session_state.new_df[st.session_state.new_df["Include"] == True],
        st.session_state.ingredient_df[st.session_state.ingredient_df["Include"] == True]
    ])

    # 선택된 부품 데이터프레임을 웹페이지에 표시
    if not selected_ingredients.empty:
        st.subheader("선택된 부품")
        st.dataframe(selected_ingredients, use_container_width=True)

    st.session_state.ingredient_df["Quantity"] = 0
    st.session_state.ingredient_df["TotalCost"] = 0

    for _, row in st.session_state.menu_df.iterrows():
        menu_id = row["ID"]
        quantity = row["Quantity"]
        menu_ingredients_df = st.session_state.menu_ingredients_df[st.session_state.menu_ingredients_df["MenuID"] == menu_id]
        for _, mi_row in menu_ingredients_df.iterrows():
            ingredient_id = mi_row["IngredientID"]
            unit_quantity = mi_row["Quantity"] * quantity

            if ingredient_id in st.session_state.ingredient_df["ID"].values:
                ingredient_idx = st.session_state.ingredient_df[st.session_state.ingredient_df["ID"] == ingredient_id].index[0]
                st.session_state.ingredient_df.at[ingredient_idx, "Quantity"] += unit_quantity
                st.session_state.ingredient_df.at[ingredient_idx, "TotalCost"] = st.session_state.ingredient_df.at[ingredient_idx, "Quantity"] * st.session_state.ingredient_df.at[ingredient_idx, "Price"]


    st.header("견적서 계산기")

    def calculate_material_costs(df):
        return df["TotalCost"].sum() if "TotalCost" in df.columns else 0

    def calculate_labor_cost(num_people):
        labor_cost_per_person = 10000
        return num_people * labor_cost_per_person

    total_material_cost = calculate_material_costs( pd.concat([
        st.session_state.new_df,
        st.session_state.ingredient_df
    ]))


    st.write(f'총 부품 비용: {total_material_cost:,} 원')

    num_people = st.number_input('인원 수 입력', min_value=0, step=1)
    total_labor_cost = calculate_labor_cost(num_people)
    st.write(f'인건비: {total_labor_cost:,} 원')

    total_cost = total_material_cost + total_labor_cost
    st.write(f'총 비용: {total_cost:,} 원')

    selected_menus_df = pd.DataFrame()
    
    for menu_name in selected_menu:
        menu_id = st.session_state.menu_df[st.session_state.menu_df["MenuName"] == menu_name]["ID"].values[0]
        quantity = st.number_input(f"{menu_name} 수량", min_value=0, step=1, key=menu_id)
        st.session_state.menu_df.loc[st.session_state.menu_df["ID"] == menu_id, "Quantity"] = quantity
        
        # 선택된 유닛를 DataFrame에 추가
        selected_menus_df = pd.concat([
            selected_menus_df,
            pd.DataFrame([[menu_id, menu_name, quantity]], columns=["ID", "MenuName", "Quantity"])
        ], ignore_index=True)

    if st.button('엑셀로 저장'):
        excel_file_path = save_to_excel(st.session_state.ingredient_df, total_material_cost, total_labor_cost, total_cost, selected_menus_df, num_people)
        st.success(f"엑셀 파일이 '{excel_file_path}'로 저장되었습니다.")

#=====================================================================================================================
# 엑셀 저장
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

def save_to_excel(materials_df, total_material_cost, labor_cost_per_person, total_cost, selected_menus_df, num_people):
    wb = Workbook()

    # 첫 번째 시트: 표제
    ws1 = wb.active
    ws1.title = "표제"

    # 스타일 설정
    thin_border = Side(border_style="thin", color="000000")
    thick_border = Side(border_style="thick", color="000000")  # 진한 선 스타일 추가
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # 표제 부분
    ws1['A1'] = "견적서"
    ws1['A1'].font = Font(size=16, bold=True)
    ws1['A1'].alignment = Alignment(horizontal="center")

    # 고객명, PROJECT, 견적금액 등의 정보 추가
    ws1['A3'] = "고객명:"
    ws1['B3'] = "고객명"
    ws1['A4'] = "PROJECT:"
    ws1['B4'] = "과제명"
    ws1['B5'] = "아래와 같이 견적을 제출 합니다."

    # 견적금액
    ws1['A7'] = "견적금액:"
    ws1['A7'].border = Border(bottom=thick_border)  # 밑줄 추가
    ws1['B7'] = f"{total_cost:,} 원"
    ws1['B7'].border = Border(bottom=thick_border)
    ws1['A8'].border = Border(bottom=thick_border)
    ws1['B8'].border = Border(bottom=thick_border)

    # 인도조건
    ws1['A9'] = "1. 인 도 조 건:"
    ws1['A9'].border = Border(bottom=thick_border)  # 밑줄 추가
    ws1['B9'] = "귀사 지정도"
    ws1['B9'].border = Border(bottom=thick_border)

    # 납기
    ws1['A10'] = "2. 납         기:"
    ws1['A10'].border = Border(bottom=thick_border) # 밑줄 추가
    ws1['B10'] = "협의"
    ws1['B10'].border = Border(bottom=thick_border)

    # 지불조건
    ws1['A11'] = "3. 지 불 조 건:"
    ws1['A11'].border = Border(bottom=thick_border)  # 밑줄 추가
    ws1['B11'] = "협의"
    ws1['B11'].border = Border(bottom=thick_border)

    # 유효기간
    ws1['A12'] = "4. 유 효 기 간:"
    ws1['A12'].border = Border(bottom=thick_border)  # 밑줄 추가
    ws1['B12'] = "견적일로 부터 30일"
    ws1['B12'].border = Border(bottom=thick_border)

    # 특기사항
    ws1['A13'] = "5. 특 기 사 항:"
    ws1['A13'].border = Border(bottom=thick_border)  # 밑줄 추가
    ws1['B13'] = ""
    ws1['B13'].border = Border(bottom=thick_border)

    # 합계금액
    ws1['A14'] = "6. 합 계 금 액:"
    ws1['A14'].border = Border(bottom=thick_border)  # 밑줄 추가
    ws1['B14'] = f"{total_cost:,} 원"
    ws1['B14'].border = Border(bottom=thick_border)

    # 담당자 정보
    ws1['C7'] = "담당자."
    ws1['D7'] = "OOO"

    ws1['C9'] = "Tel. 041-"
    ws1['D9'] = ""
    ws1['C10'] = "Fax. 041-"
    ws1['D10'] = ""
    ws1['C11'] = "Mobile. 010-"
    ws1['D11'] = ""
    ws1['C12'] = "Quotation No."
    ws1['D12'] = ""
    ws1['C13'] = "Quotation Date."
    ws1['D13'] = ""
    ws1['C14'] = "담당자."
    ws1['D14'] = "사업자번호."

    # 열 너비 조정
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 6
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 20

    # 유닛 정보 표 헤더
    headers = ['번호', '유닛명', '수량', '단위', '단가', '합계', '비고']

    # 표제 시트에서 16번째 행에 헤더 추가
    start_row = 16

    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=start_row, column=col_idx, value=header)
        cell.border = thin_border
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # 유닛 정보 추가
    for idx, row in selected_menus_df.iterrows():
        menu_id = row['ID']
        menu_name = row['MenuName']
        quantity = row['Quantity']
        unit = '단위'  # 단위를 지정해 주세요
        price = st.session_state.menu_df[st.session_state.menu_df["ID"] == menu_id]["TotalCost"].values[0] / quantity
        total_cost = price * quantity
        ws1.append([
            idx + 1,
            menu_name,
            quantity,
            unit,
            price,
            total_cost,
            ""  # 비고는 빈칸
        ])

    # 인건비 추가
    labor_total_cost = labor_cost_per_person * num_people
    ws1.append([
        len(selected_menus_df) + 1,
        '인건비',
        num_people,
        '명',
        labor_cost_per_person,
        labor_total_cost,
        ""
    ])

    # 모든 셀에 테두리 추가
    for row in ws1.iter_rows(min_row=start_row, max_row=ws1.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
            if cell.column in [5, 6]:  # '단가 (KRW)'와 '합계 (KRW)' 컬럼
                cell.number_format = '#,##0 원'

    # 가운데 정렬
    for row in range(3, 100):  # A3부터 A19까지
        cell = ws1[f'A{row}']
        cell.alignment = center_align

    for row in range(17, 100):  # C17부터 C19까지
        cell = ws1[f'C{row}']
        cell.alignment = center_align

    for row in range(17, 100):  # D17부터 D19까지
        cell = ws1[f'D{row}']
        cell.alignment = center_align

    #---------------------------------------------------------
    # 두 번째 페이지: 세부 부품 및 비용 정보
    ws2 = wb.create_sheet(title="세부")

    # 세부 정보 표 헤더
    ws2.append(['번호', '부품명', '수량', '단위', '단가', '합계'])

    for cell in ws2[1]:
        cell.border = thin_border
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # 세부 부품 및 비용 정보 추가
    for idx, row in materials_df.iterrows():
        ws2.append([row['ID'], row['IngredientName'], row['Quantity'], '단위', row['Price'], row['TotalCost']])

    # 모든 셀에 테두리 추가
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
            if cell.column in [5, 6]:  # '단가 (KRW)'와 '합계 (KRW)' 컬럼
                cell.number_format = '#,##0 원'

    # 세부 정보 시트에서 합계 추가
    ws2.append(['', '', '', '', '부품 총 비용:', f"{total_material_cost:,} 원"])
    ws2.append(['', '', '', '', '인건비:', f"{labor_total_cost:,} 원"])
    ws2.append(['', '', '', '', '전체 총 비용:', f"{total_material_cost + labor_total_cost:,} 원"])

    # 가운데 정렬
    for row in range(1, 100): 
        cell = ws2[f'A{row}']
        cell.alignment = center_align

    for row in range(1, 100): 
        cell = ws2[f'B{row}']
        cell.alignment = center_align

    for row in range(1, 100): 
        cell = ws2[f'C{row}']
        cell.alignment = center_align

    for row in range(1, 100): 
        cell = ws2[f'D{row}']
        cell.alignment = center_align

    for row in range(21, 24):  # E21부터 E23까지
        cell = ws2[f'E{row}']
        cell.alignment = center_align

    # 열 너비 조정
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 15

    excel_file_path = '견적서.xlsx'
    wb.save(excel_file_path)
    return excel_file_path

if __name__ == "__main__":
    main()